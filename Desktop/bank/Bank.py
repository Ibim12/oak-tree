import email
from time import time
from unicodedata import name
from Current import Current
from FixedDep import Fixed_Deposit
from Savings import Savings
from User import User_Account
from random import randint
from Acno import Account_number
from random import randint
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import date

wb= Workbook()
ws = wb.active
ws.title = "Book1"

today= date.today()
import datetime
now = datetime.datetime.now()



class Bank:
    total_user=[]
    banks_money=0
    def __init__(self,username,account_number,account_type) -> None:
        self.username= username
        self.account_number= ''
        self.set_user_account_number()
        self.interest_rate= ""
        self.savings_account= Savings()
        self.current_account= Current()
        self.fixed_deposit_account= Fixed_Deposit()
        
    def __init__(self) -> None:
        pass

    def receive_deposit(self):
        pass
    def give_money(self,account_number,amount):
        for user in self.total_user:
            if user.get_account_number()==account_number:
                current_balance=user.get_account_balance()-amount
                user.set_account_balance(current_balance)
    
    def get_account_number(self):
        return self.account_number
    def create_account_number(self):       
        return Account_number.account_number_generation(self)
            

    def give_account_type(self):
        pass
    def create_user(self,firstname,lastname,email,account_balance,password,account_type):
        user=User_Account(firstname,lastname,email,account_balance,password,account_type)
        user.set_user_account_number(self.create_account_number())
        self.total_user.append(user)
    
    def display_all_users(self):
        for user in self.total_user:
            print(user)
    def get_all_user_details(self):
        return User_Account.get_user_details()

        




bank= Bank()
bank.create_user( 'Lilah','Jakande','lj@gmail.com',109000,'21344', 'Current')

# print(bank.get_all_user_details)
# print(bank.create_account_number())

bank.create_user( 'Tamal','Abdulfitri','TA@gmail.com',234000,'89253', 'Savings')
#bank.get_account_number.bank.total_user[0]
bank.give_money('973636490',1000)
bank.create_user('Theresa', 'Belema','TB@gmail.com', 229200,'12723', 'FixedDep')
# print(bank.total_user[0])
bank.display_all_users()
# bank.get_all_user_details()
#print(Bank.collect_user_details)

# for i in range(1,5):
#     def db_structure(self,num):
#       self.num=1 
#       ws['Anum'].value= firstname
#       ws['Bnum'].value= lastname
#       ws['Cnum'].value= Account_number
#       ws['Dnum'].value= account_type
#       ws['Enum'].value= email
#       ws['Fnum'].value=date
#       ws['Gnum'].value= time 
#         def __init__(self,firstname,lastname,account_number,account_type):
#             pass
    
        
ws['A1'].value ='firstname'
ws['B1'].value ='lastname'
ws['C1'].value= 'account balance'
ws['D1'].value= 'account number'
ws['E1'].value= 'email'
ws['F1'].value= 'date'
ws['G1'].value= 'time'

ws['A2'].value= bank.total_user[0].get_firstname()
ws['B2'].value= bank.total_user[0].get_lastname()
ws['C2'].value= bank.total_user[0].get_account_balance()
ws['D2'].value= bank.total_user[0].get_account_number()
ws['E2'].value= bank.total_user[0].get_email()
ws['F2'].value= today.strftime("%d/%m/%y")
ws['G2'].value= today.strftime("%H/%M/%S")

ws['A3'].value= bank.total_user[1].get_firstname()
ws['B3'].value= bank.total_user[1].get_lastname()
ws['C3'].value= bank.total_user[1].get_account_balance()
ws['D3'].value= bank.total_user[1].get_account_number()
ws['E3'].value= bank.total_user[1].get_email()
ws['F3'].value= today.strftime("%d/%m/%y")
ws['G3'].value= today.strftime("%H/%M/%S")

ws['A4'].value= bank.total_user[2].get_firstname()
ws['B4'].value= bank.total_user[2].get_lastname()
ws['C4'].value= bank.total_user[2].get_account_balance()
ws['D4'].value= bank.total_user[2].get_account_number()
ws['E4'].value= bank.total_user[2].get_email()
ws['F4'].value= today.strftime("%d/%m/%y")
ws['G4'].value= today.strftime("%H/%M/%S")

wb.save('Book1.xlsx')
