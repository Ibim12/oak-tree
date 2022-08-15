from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Bank import Bank
wb= Workbook()
ws = wb.active
ws.title = "Bank"

# ws.append(Bank.total_user)
# for row in range(1,10):
    
for col in range(1,5):
    char = get_column_letter(col)
    ws[char + str(row)]= 'char'
for user in Bank.display_all_users():
    data = list([user].values())
    ws.append([user]+data)

wb.save('Bank.xlsx')