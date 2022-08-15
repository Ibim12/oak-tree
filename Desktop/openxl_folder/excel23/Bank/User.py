from random import randint
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb= Workbook()
ws = wb.active
ws.title = "Book1"

# ws.append(Bank.total_user)
# for row in range(1,10):
    




class User_Account:
    def __init__(self,firstname,lastname,email,account_balance,password,account_type)-> None:
        self.firstname = firstname
        self.lastname = lastname
        self.email = email
        self.account_balance = account_balance 
        self.password = password
        self.account_type= account_type
        self.account_number= ''
        

    def get_firstname(self):
        return self.firstname

    def set_firstname(self, firstname):
        self._firstname = firstname

    def get_lastname(self):
        return self.lastname

    def set_lastname(self, lastname):
        self._lastname = lastname

    def get_email(self):
        return self.email

    def set_email(self, email):
        self._email = email

    def get_account_balance(self):
        return self.account_balance

    def set_account_balance(self, account_balance):
        self.account_balance = account_balance

    def get_password(self):
        return self.password

    def set_password(self, password):
        self._password = password
    
    def get_account_type(self):
        return self.account_type

    def set_(self, account_type):
        self._account_type = account_type

    def get_transaction_history(self):
        return self.transaction_history

    def set_(self, transaction_history):
        self._transaction_history= transaction_history

    def get_account_number(self):
        return self.account_number

    def set_user_account_number(self,account_number):
        self.account_number = account_number
    def get_user_details(self):
        return User_Account()
    def __str__(self):
        return f"name:{self.get_firstname()}-{self.get_lastname()}  account_balance:{self.get_account_balance() } account_number :{self.get_account_number()} email:{self.get_email()}"
user=User_Account('Lilah','Jakande','lj@gmail.com',109000,'21344', 'Current')
print(user)
# print(user)
# user.set_user_account_number()
# print(user.get_account_number())

# user_string=str(user)
# user_list= list(user_string)
# print(user_list)
# user_entry=''.join(str(e) for e in user_list)
# print(user_entry)
# # for col in range(1,5):
#     char = get_column_letter(col)
#     ws[char + str(row)]= 'char'
# for user in Bank.display_all_users():
#     data = list([user].values())
#     ws.append([user]+data)
# user_entry_1 =set(user_entry)
# # ws.append(user_entry_1)
# print(user_entry_1)


ws['A1'].value ='firstname'
ws['B1'].value ='lastname'
ws['C1'].value= 'account balance'
ws['D1'].value= 'account number'
ws['E1'].value= 'email'
ws['A2'].value= user.get_firstname()
ws['B2'].value= user.get_lastname()
ws['C2'].value= user.get_account_balance()
ws['D2'].value= user.get_account_number()
ws['E2'].value= user.get_email()
wb.save('Book1.xlsx')
