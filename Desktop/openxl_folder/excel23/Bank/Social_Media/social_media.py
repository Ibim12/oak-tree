from social_media_user import social_media_user_account
from social_media_posts_generation import Social_Media_Posts_generation
from openpyxl import workbook, load_workbook
from datetime import datetime
from datetime import date

wb = load_workbook(r"C:\Users\IBIM\Desktop\openxl_folder\Bank.xlsx")
ws= wb.active
today= date.today()
class Social_Media:
    def __init__(self,name,Social_media_user,social_media_posts):
        self.name = name
        # self.Social_media_user = Social_Media_Posts_generation.get_user_account()
        self.social_media_posts= social_media_posts
        self.Social_media_posts_generation= Social_Media_Posts_generation(name, 'user_account','Social_Media_Post')
    # def db_structure(self,num,name,Social_media_user,social_media_posts,date,time):
        # date= today.strftime("%d/%m/%y")
        # time= today.strftime("%H/%M/%S")
        # for i in range(1,5):
            # self.num=1 
soc=Social_Media('william','willylidi','picture2')      
print(soc)      
ws['A1'].value= Social_Media_Posts_generation.get_name()
ws['B1'].value= soc.Social_Media_Posts_generation.get_user_account()
ws['C1'].value= soc.social_Media_Posts_generation.get_Social_Media_Post()
ws['D1'].value= today.strftime("%d/%m/%y")
ws['E1'].value= today.strftime("%H/%M/%S")
            # self.num+=1

wb.save(r"C:\Users\IBIM\Desktop\openxl_folder\Bank.xlsx")