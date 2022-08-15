from tkinter import font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb = load_workbook("C:\\Users\\IBIM\\Documents\\Assessments.xlsx")
ws=wb.active
for row in range(1,10):
    for col in range(1,5):
            char = get_column_letter(col)
            ws[char + str(row)]= char + str(row)
    ws.title='Paper'
    ws.merge_cells=("A1:D1")
    ws.insert_cols(2)
    ws.delete_cols(3)
    ws.move_range("C1:D11", rows=13, cols= 9)